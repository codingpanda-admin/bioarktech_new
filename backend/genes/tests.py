from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook, load_workbook

from products.models import Product

from .models import GeneDesignCategory, GeneDesignPrice, GeneLibrary
from .bulk_upsert import GENE_HEADERS, build_gene_upsert_template, import_gene_upsert_workbook
from .pricing import build_design_sku, resolve_design_sku_price
from .views import _normalize_target_gene_code


class GeneDesignMetadataTests(TestCase):
    def test_metadata_endpoint_returns_all_six_design_steps(self):
        response = self.client.get('/api/genes/design-metadata/')

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(
            set(data),
            {
                'categories',
                'delivery_types',
                'structure_substeps',
                'target_gene_options',
                'format_types',
            },
        )
        self.assertEqual(len(data['categories']), 4)
        self.assertEqual(len(data['delivery_types']), 4)
        self.assertEqual(len(data['structure_substeps']), 6)
        self.assertEqual(len(data['target_gene_options']), 2)
        self.assertEqual(len(data['format_types']), 3)

        crispr = next(
            category
            for category in data['categories']
            if category['name'] == 'CRISPR-Cas9'
        )
        self.assertEqual(len(crispr['function_types']), 5)

        format_option_count = sum(
            len(format_type['options']) for format_type in data['format_types']
        )
        self.assertEqual(format_option_count, 5)

    def test_metadata_endpoint_excludes_inactive_values(self):
        GeneDesignCategory.objects.create(
            code='inactive-test',
            name='Inactive Test Category',
            description='This category must not be exposed.',
            is_active=False,
        )

        response = self.client.get('/api/genes/design-metadata/')

        self.assertEqual(response.status_code, 200)
        names = [item['name'] for item in response.json()['categories']]
        self.assertNotIn('Inactive Test Category', names)


class GeneDesignPriceLookupTests(TestCase):
    def setUp(self):
        self.base_payload = {
            'function_type_code': 'CA',
            'delivery_type_code': 'S',
            'target_gene_code': '000000',
            'structure_map': {
                'S1': '0',
                'S2': 'X',
                'S3': '0',
                'S4': '0',
                'S5': '0',
                'S6': 'A',
            },
            'formats': [{'code_id': 'k', 'unit_amount': '5ug DNA'}],
        }

    def post_lookup(self, payload):
        return self.client.post(
            '/api/genes/design-price/',
            payload,
            content_type='application/json',
        )

    def test_final_sku_uses_step_codes_with_format_separator(self):
        self.assertEqual(
            build_design_sku(self.base_payload, 'k'),
            'CAS-0X000A-000000-k',
        )

    def test_seeded_reference_contains_80_unique_rules(self):
        self.assertEqual(GeneDesignPrice.objects.count(), 80)

    def test_non_cd_uses_others_custom_price(self):
        response = self.post_lookup(self.base_payload)

        self.assertEqual(response.status_code, 200)
        price = response.json()['results'][0]
        self.assertEqual(price['status'], 'priced')
        self.assertEqual(price['currency'], 'USD')
        self.assertEqual(price['list_price'], '399.00')
        self.assertEqual(price['discount_price'], '198.50')
        self.assertNotIn('shelf_status', price)

    def test_matching_catalog_sku_uses_backend_shelf_price(self):
        Product.objects.create(
            external_id='shelf-test-product',
            product_name='Shelf Test Product',
            catalog_number='CAS-0X000A-000000-k',
            hidden=False,
        )

        response = self.post_lookup(self.base_payload)

        price = response.json()['results'][0]
        self.assertEqual(price['list_price'], '249.00')
        self.assertEqual(price['discount_price'], '123.50')

    def test_specific_gene_is_normalized_to_hash_bucket(self):
        payload = {
            **self.base_payload,
            'target_gene_code': 'ABC123',
            'formats': [{'code_id': 'k', 'unit_amount': '5ug each/3 tubes plus control'}],
        }

        response = self.post_lookup(payload)

        price = response.json()['results'][0]
        self.assertEqual(_normalize_target_gene_code('ABC123'), '######')
        self.assertEqual(price['list_price'], '799.00')
        self.assertEqual(price['discount_price'], '398.50')

    def test_cd_lentivirus_ignores_step_five_and_submits_quote(self):
        payload = {
            **self.base_payload,
            'function_type_code': 'CD',
            'delivery_type_code': 'L',
            'target_gene_code': 'ANY123',
            'formats': [{'code_id': 'l', 'unit_amount': '1X10^7 IU/ml'}],
        }

        response = self.post_lookup(payload)

        price = response.json()['results'][0]
        self.assertEqual(price['status'], 'quote_only')
        self.assertTrue(price['quote_only'])
        self.assertIsNone(price['list_price'])
        self.assertIsNone(price['discount_price'])
        self.assertEqual(price['action'], 'Submit Quote')

    def test_multiple_format_buckets_return_separate_prices(self):
        payload = {
            **self.base_payload,
            'target_gene_code': 'ABC123',
            'formats': [
                {'code_id': 'k', 'unit_amount': '5ug each/3 tubes plus control'},
                {'code_id': 'c', 'unit_amount': '1X10^6 Cells'},
            ],
        }

        response = self.post_lookup(payload)

        results = response.json()['results']
        self.assertEqual(response.status_code, 200)
        self.assertEqual([item['format_code'] for item in results], ['k', 'c'])
        self.assertEqual(results[0]['discount_price'], '398.50')
        self.assertEqual(results[1]['discount_price'], '4520.00')

    def test_cart_price_is_resolved_from_the_backend_rule(self):
        resolved = resolve_design_sku_price(
            'CAS-0X000A-000000-k',
            '5ug DNA',
        )

        self.assertTrue(resolved['recognized'])
        self.assertFalse(resolved['quote_only'])
        self.assertEqual(str(resolved['price']), '198.50')

    def test_legacy_cart_sku_remains_price_compatible(self):
        resolved = resolve_design_sku_price(
            'CAS-0X000A-000000k',
            '5ug DNA',
        )

        self.assertTrue(resolved['recognized'])
        self.assertFalse(resolved['quote_only'])
        self.assertEqual(str(resolved['price']), '198.50')

    def test_quote_only_design_cannot_resolve_a_cart_price(self):
        resolved = resolve_design_sku_price(
            'CDL-0X000C-ANY123-l',
            '1X10^7 IU/ml',
        )

        self.assertTrue(resolved['recognized'])
        self.assertTrue(resolved['quote_only'])
        self.assertIsNone(resolved['price'])


class GeneLibrarySearchTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        GeneLibrary.objects.bulk_create([
            GeneLibrary(
                target_sequence='HGENE1',
                gene_name='Human kinase alpha',
                symbol='HKA',
                species='Human',
                description='Regulates cell growth signaling.',
            ),
            GeneLibrary(
                target_sequence='MGENE1',
                gene_name='Mouse phosphatase beta',
                symbol='MPB',
                species='Mouse',
                description='Controls neural differentiation.',
            ),
        ])

    def test_searches_name_and_filters_species(self):
        response = self.client.get(
            '/api/genes/gene-library/',
            {'species': 'Human', 'gene_name': 'kinase'},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['total'], 1)
        self.assertEqual(data['results'][0]['target_sequence'], 'HGENE1')

    def test_searches_description_without_description_filter_values(self):
        response = self.client.get(
            '/api/genes/gene-library/',
            {'description': 'neural'},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['total'], 1)
        self.assertEqual(data['results'][0]['target_sequence'], 'MGENE1')

    def test_finds_gene_by_exact_target_sequence_case_insensitively(self):
        response = self.client.get(
            '/api/genes/gene-library/',
            {'target_sequence': 'hgene1'},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['total'], 1)
        self.assertEqual(data['results'][0]['target_sequence'], 'HGENE1')


class GeneLibraryBulkUpsertTests(TestCase):
    @staticmethod
    def workbook(rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Upload'
        sheet.append(GENE_HEADERS)
        for row in rows:
            sheet.append([row.get(header, '') for header in GENE_HEADERS])
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def test_template_contains_expected_upload_columns(self):
        template = build_gene_upsert_template()
        workbook = load_workbook(template, read_only=True)
        self.assertEqual(workbook.sheetnames, ['Upload', 'Instructions'])
        headers = list(next(workbook['Upload'].iter_rows(values_only=True)))
        self.assertEqual(headers, GENE_HEADERS)
        workbook.close()

    def test_upsert_matches_target_sequence_case_insensitively(self):
        existing = GeneLibrary.objects.create(
            target_sequence='AbC123',
            gene_name='Old Gene Name',
            abbreviation='OLD',
            symbol='OLD',
            species='Human',
        )
        upload = self.workbook([
            {
                'target_sequence': 'abc123',
                'gene_name': 'Updated Gene Name',
                'symbol': 'NEW',
                'species': 'Mouse',
            },
            {
                'target_sequence': 'DEF456',
                'gene_name': 'Created Gene',
                'symbol': 'CG',
                'locus_id': 12345,
            },
        ])

        result = import_gene_upsert_workbook(upload)

        self.assertEqual(result['created'], 1)
        self.assertEqual(result['updated'], 1)
        self.assertEqual(result['failed'], 0)
        existing.refresh_from_db()
        self.assertEqual(existing.target_sequence, 'ABC123')
        self.assertEqual(existing.gene_name, 'Updated Gene Name')
        self.assertEqual(existing.symbol, 'NEW')
        self.assertEqual(existing.species, 'Mouse')
        self.assertIsNone(existing.abbreviation)
        self.assertTrue(GeneLibrary.objects.filter(target_sequence='DEF456', locus_id=12345).exists())

    def test_duplicate_target_sequences_in_workbook_are_rejected(self):
        upload = self.workbook([
            {'target_sequence': 'ABC123', 'gene_name': 'Gene One', 'symbol': 'ONE'},
            {'target_sequence': 'abc123', 'gene_name': 'Gene Two', 'symbol': 'TWO'},
        ])

        result = import_gene_upsert_workbook(upload)

        self.assertEqual(result['created'], 0)
        self.assertEqual(result['updated'], 0)
        self.assertEqual(result['failed'], 2)
        self.assertEqual(GeneLibrary.objects.filter(target_sequence__iexact='ABC123').count(), 0)
