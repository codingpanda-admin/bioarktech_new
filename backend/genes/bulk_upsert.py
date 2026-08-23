"""Excel template and validated bulk upsert support for the Gene Library."""

from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.db import transaction
from django.db.models.functions import Upper
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import GeneLibrary


GENE_HEADERS = [
    'target_sequence',
    'gene_name',
    'symbol',
    'abbreviation',
    'locus_id',
    'species',
    'description',
    'reference_link',
]
REQUIRED_HEADERS = {'target_sequence', 'gene_name', 'symbol'}


def _cell_text(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return format(value, 'f')
    return str(value).strip()


def _normalized_header(value):
    return _cell_text(value).lower().replace(' ', '_').replace('-', '_')


def _locus_id(value):
    text = _cell_text(value)
    if not text:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError('locus_id must be a whole number.') from exc
    if not number.is_finite() or number != number.to_integral_value():
        raise ValueError('locus_id must be a whole number.')
    return int(number)


def _gene_values(row):
    target_sequence = _cell_text(row.get('target_sequence')).upper()
    gene_name = _cell_text(row.get('gene_name'))
    symbol = _cell_text(row.get('symbol'))
    if not target_sequence:
        raise ValueError('target_sequence is required.')
    if len(target_sequence) > 6:
        raise ValueError('target_sequence cannot exceed 6 characters.')
    if not gene_name:
        raise ValueError('gene_name is required.')
    if not symbol:
        raise ValueError('symbol is required.')

    return {
        'target_sequence': target_sequence,
        'gene_name': gene_name,
        'symbol': symbol,
        'abbreviation': _cell_text(row.get('abbreviation')) or None,
        'locus_id': _locus_id(row.get('locus_id')),
        'species': _cell_text(row.get('species')) or None,
        'description': _cell_text(row.get('description')) or None,
        'reference_link': _cell_text(row.get('reference_link')) or None,
    }


def build_gene_upsert_template():
    """Return an in-memory workbook for Gene Library bulk upserts."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Upload'
    sheet.append(GENE_HEADERS)
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:H1'

    header_fill = PatternFill('solid', fgColor='1D4ED8')
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 24
    widths = [19, 42, 20, 20, 16, 18, 55, 45]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    # Preserve leading zeroes and alphanumeric target codes.
    sheet.column_dimensions['A'].number_format = '@'

    instructions = workbook.create_sheet('Instructions')
    instructions['A1'] = 'BioArk Gene Library Bulk Upsert'
    instructions['A1'].font = Font(size=16, bold=True, color='17365D')
    instructions['A3'] = 'Rule'
    instructions['B3'] = 'Details'
    for cell in instructions[3]:
        cell.fill = PatternFill('solid', fgColor='DCE6F1')
        cell.font = Font(bold=True)
    rules = [
        ('One gene per row', 'Enter records on the Upload sheet beginning on row 2.'),
        ('Upsert key', 'target_sequence is matched case-insensitively. Existing matches are updated; new values are created.'),
        ('Required fields', 'target_sequence, gene_name, and symbol are required on every populated row.'),
        ('Target sequence', 'Maximum 6 characters. Keep this column formatted as Text to preserve leading zeroes.'),
        ('Optional values', 'Leave an optional cell blank to clear that value when an existing record is updated.'),
        ('Duplicate rows', 'A target_sequence may appear only once in a workbook.'),
    ]
    for rule in rules:
        instructions.append(rule)
    instructions.column_dimensions['A'].width = 24
    instructions.column_dimensions['B'].width = 105
    for row in instructions.iter_rows(min_row=4, max_row=3 + len(rules), min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _existing_genes_by_sequence(sequences):
    matches = defaultdict(list)
    sequence_list = sorted(set(sequences))
    for start in range(0, len(sequence_list), 500):
        chunk = sequence_list[start:start + 500]
        genes = GeneLibrary.objects.annotate(
            normalized_target_sequence=Upper('target_sequence'),
        ).filter(normalized_target_sequence__in=chunk)
        for gene in genes:
            matches[str(gene.target_sequence or '').strip().upper()].append(gene)
    return matches


def import_gene_upsert_workbook(uploaded_file):
    """Create or update Gene Library rows, using target_sequence as the key."""
    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('The uploaded file is not a readable Excel workbook.') from exc

    try:
        if 'Upload' not in workbook.sheetnames:
            raise ValueError('The workbook must contain a sheet named "Upload".')
        sheet = workbook['Upload']
        raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [_normalized_header(value) for value in raw_headers]
        nonblank_headers = [header for header in headers if header]
        if len(nonblank_headers) != len(set(nonblank_headers)):
            raise ValueError('The Upload sheet contains duplicate column headers.')

        missing = sorted(REQUIRED_HEADERS - set(nonblank_headers))
        if missing:
            raise ValueError(f'Missing required columns: {", ".join(missing)}.')
        unknown = sorted(set(nonblank_headers) - set(GENE_HEADERS))
        if unknown:
            raise ValueError(f'Unknown columns: {", ".join(unknown)}.')

        parsed_rows = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row = {header: value for header, value in zip(headers, values) if header}
            if any(_cell_text(value) for value in row.values()):
                parsed_rows.append((row_number, row))
    finally:
        workbook.close()

    if not parsed_rows:
        raise ValueError('The Upload sheet does not contain any data rows.')

    result = {
        'total_rows': len(parsed_rows),
        'created': 0,
        'updated': 0,
        'failed': 0,
        'errors': [],
    }
    validated = []
    for row_number, row in parsed_rows:
        try:
            validated.append((row_number, _gene_values(row)))
        except ValueError as exc:
            result['failed'] += 1
            result['errors'].append({
                'row': row_number,
                'target_sequence': _cell_text(row.get('target_sequence')).upper(),
                'message': str(exc),
            })

    workbook_counts = Counter(values['target_sequence'] for _, values in validated)
    duplicate_sequences = {sequence for sequence, count in workbook_counts.items() if count > 1}
    unique_validated = []
    for row_number, values in validated:
        if values['target_sequence'] in duplicate_sequences:
            result['failed'] += 1
            result['errors'].append({
                'row': row_number,
                'target_sequence': values['target_sequence'],
                'message': 'This target_sequence appears more than once in the workbook.',
            })
        else:
            unique_validated.append((row_number, values))

    existing = _existing_genes_by_sequence(
        values['target_sequence'] for _, values in unique_validated
    )
    to_create = []
    to_update = []
    update_fields = list(GENE_HEADERS)
    for row_number, values in unique_validated:
        matches = existing.get(values['target_sequence'], [])
        if len(matches) > 1:
            result['failed'] += 1
            result['errors'].append({
                'row': row_number,
                'target_sequence': values['target_sequence'],
                'message': 'Multiple database records use this target_sequence; resolve the duplicates before upserting.',
            })
            continue
        if matches:
            gene = matches[0]
            for field, value in values.items():
                setattr(gene, field, value)
            to_update.append(gene)
        else:
            to_create.append(GeneLibrary(**values))

    with transaction.atomic():
        if to_create:
            GeneLibrary.objects.bulk_create(to_create, batch_size=1000)
        if to_update:
            GeneLibrary.objects.bulk_update(to_update, update_fields, batch_size=1000)

    result['created'] = len(to_create)
    result['updated'] = len(to_update)
    return result
