"""Validated Excel bulk imports for the public catalog.

The import intentionally excludes media and document fields. Rich-text-capable
model fields receive the cell text verbatim, without applying HTML formatting.
"""

from decimal import Decimal, InvalidOperation

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from interface.models import ServiceMode
from products.models import CatalogGroup, Product, ProductCategory


PRODUCT_HEADERS = [
    'external_id',
    'product_name',
    'category_external_id',
    'group_external_id',
    'catalog_number',
    'show_catalog_number',
    'description',
    'details',
    'key_features',
    'storage_stability',
    'performance_data',
    'list_price',
    'discounted_price',
    'first_option_name',
    'first_option_list_price',
    'first_option_discounted_price',
    'availability',
    'price_range',
    'quote_only',
    'featured',
    'display_on_homepage',
    'active',
    'display_order',
]

SERVICE_HEADERS = [
    'external_id',
    'service_name',
    'category_external_id',
    'group_external_id',
    'catalog_number',
    'show_catalog_number',
    'service_details',
    'technique',
    'price',
    'performance_data',
    'featured',
    'recommended_service',
    'display_on_homepage',
    'active',
]

REQUIRED_HEADERS = {
    'product': {'external_id', 'product_name', 'category_external_id'},
    'reagent': {'external_id', 'product_name', 'category_external_id'},
    'service': {'external_id', 'service_name', 'category_external_id'},
}


def _cell_text(value):
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, Decimal):
        return format(value, 'f')
    return str(value).strip()


def _normalized_header(value):
    return _cell_text(value).lower().replace(' ', '_').replace('-', '_')


def _parse_bool(value, field_name, default=False):
    text = _cell_text(value).lower()
    if not text:
        return default
    if text in {'yes', 'true', '1', 'y'}:
        return True
    if text in {'no', 'false', '0', 'n'}:
        return False
    raise ValueError(f'{field_name} must be Yes or No.')


def _parse_integer(value, field_name):
    text = _cell_text(value)
    if not text:
        return None
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f'{field_name} must be a whole number.') from exc
    if number != number.to_integral_value():
        raise ValueError(f'{field_name} must be a whole number.')
    return int(number)


def _split_lines(value):
    return [line.strip() for line in _cell_text(value).splitlines() if line.strip()]


def _error_text(exc):
    if isinstance(exc, ValidationError):
        if hasattr(exc, 'message_dict'):
            return '; '.join(
                f'{field}: {", ".join(str(message) for message in messages)}'
                for field, messages in exc.message_dict.items()
            )
        return '; '.join(str(message) for message in exc.messages)
    return str(exc)


def _category_for(external_id, item_type):
    category_id = _cell_text(external_id)
    matches = ProductCategory.objects.filter(external_id=category_id)
    if not category_id or not matches.exists():
        raise ValueError(f'Unknown category_external_id "{category_id}".')
    if matches.count() > 1:
        raise ValueError(f'category_external_id "{category_id}" is not unique.')
    category = matches.first()
    category_type = _cell_text(category.product_type).lower()
    allowed = {
        'product': {'product', 'both'},
        'reagent': {'reagent', 'consumable'},
        'service': {'service'},
    }[item_type]
    if category_type and category_type not in allowed:
        raise ValueError(
            f'Category "{category.category_name}" cannot contain a {item_type} item.'
        )
    return category


def _group_for(external_id, category):
    group_id = _cell_text(external_id)
    if not group_id:
        return None
    group = CatalogGroup.objects.filter(external_id=group_id).first()
    if not group:
        raise ValueError(f'Unknown group_external_id "{group_id}".')
    if group.category_id != category.category_id:
        raise ValueError(
            f'Group "{group.group_name}" does not belong to category "{category.category_name}".'
        )
    if not group.is_active:
        raise ValueError(f'Group "{group.group_name}" is inactive.')
    return group


def _find_product(external_id):
    matches = Product.objects.filter(external_id__iexact=external_id)
    if matches.count() > 1:
        raise ValueError(f'External ID "{external_id}" matches multiple existing items.')
    return matches.first()


def _import_product(row, item_type):
    external_id = _cell_text(row.get('external_id'))
    name = _cell_text(row.get('product_name'))
    if not external_id:
        raise ValueError('external_id is required.')
    if not name:
        raise ValueError('product_name is required.')

    category = _category_for(row.get('category_external_id'), item_type)
    group = _group_for(row.get('group_external_id'), category)
    product = _find_product(external_id)
    created = product is None
    if created:
        product = Product(external_id=external_id)
    else:
        existing_type = (
            'reagent'
            if _cell_text(product.source_type).lower() == 'reagent'
            or _cell_text(getattr(product.category, 'product_type', '')).lower() in {'reagent', 'consumable'}
            else 'product'
        )
        if existing_type != item_type:
            raise ValueError(
                f'External ID "{external_id}" already belongs to a {existing_type}.'
            )

    option_name = _cell_text(row.get('first_option_name'))
    option_price = _cell_text(row.get('first_option_list_price'))
    option_discount = _cell_text(row.get('first_option_discounted_price'))
    if (option_price or option_discount) and not option_name:
        raise ValueError('first_option_name is required when an option price is provided.')

    product.external_id = external_id
    product.product_name = name
    product.category = category
    product.category_external_id = category.external_id
    product.catalog_group = group
    product.product_group = group.group_name if group else ''
    product.source_type = 'reagent' if item_type == 'reagent' else 'quote'
    product.catalog_number = _cell_text(row.get('catalog_number'))
    product.show_catalog_number = _parse_bool(
        row.get('show_catalog_number'), 'show_catalog_number', default=True
    )
    product.description = _cell_text(row.get('description'))
    product.content_text = _cell_text(row.get('details'))
    product.key_features = _split_lines(row.get('key_features'))
    product.storage_stability = _cell_text(row.get('storage_stability'))
    product.performance_data = _cell_text(row.get('performance_data'))
    product.list_price = _cell_text(row.get('list_price'))
    product.discounted_price = _cell_text(row.get('discounted_price'))
    product.options = [option_name] if option_name else []
    product.option_prices = {option_name: option_price} if option_name and option_price else {}
    product.option_discounted_prices = (
        {option_name: option_discount} if option_name and option_discount else {}
    )
    product.availability = _cell_text(row.get('availability'))
    product.price_range = _cell_text(row.get('price_range'))
    product.quote_only = _parse_bool(row.get('quote_only'), 'quote_only', default=False)
    product.is_featured = _parse_bool(row.get('featured'), 'featured', default=False)
    product.show_on_screen = _parse_bool(
        row.get('display_on_homepage'), 'display_on_homepage', default=False
    )
    product.hidden = not _parse_bool(row.get('active'), 'active', default=True)
    product.display_order = _parse_integer(row.get('display_order'), 'display_order')
    product.save()
    return created, product.external_id, product.product_name


def _find_service(external_id):
    matches = ServiceMode.objects.filter(url__iexact=external_id)
    if matches.count() > 1:
        raise ValueError(f'External ID "{external_id}" matches multiple existing services.')
    return matches.first()


def _import_service(row):
    external_id = _cell_text(row.get('external_id'))
    name = _cell_text(row.get('service_name'))
    if not external_id:
        raise ValueError('external_id is required.')
    if not name:
        raise ValueError('service_name is required.')
    if len(name) > 60:
        raise ValueError('service_name cannot exceed 60 characters.')

    category = _category_for(row.get('category_external_id'), 'service')
    group = _group_for(row.get('group_external_id'), category)
    service = _find_service(external_id)
    created = service is None
    if created:
        service = ServiceMode(url=external_id, title=name, content='')

    service.url = external_id
    service.title = name
    service.category_ref = category
    service.category = category.external_id
    service.catalog_group = group
    service.service_group = group.group_name if group else ''
    service.catalog_number = _cell_text(row.get('catalog_number'))
    service.show_catalog_number = _parse_bool(
        row.get('show_catalog_number'), 'show_catalog_number', default=True
    )
    service.content = _cell_text(row.get('service_details'))
    service.technique = _cell_text(row.get('technique'))
    service.price = _cell_text(row.get('price'))
    service.performance_data = _cell_text(row.get('performance_data'))
    service.is_featured = _parse_bool(row.get('featured'), 'featured', default=False)
    service.presented_service = _parse_bool(
        row.get('recommended_service'), 'recommended_service', default=False
    )
    service.show_on_screen = _parse_bool(
        row.get('display_on_homepage'), 'display_on_homepage', default=False
    )
    service.hidden = not _parse_bool(row.get('active'), 'active', default=True)
    service.save()
    return created, service.url, service.title


def import_catalog_workbook(uploaded_file, item_type):
    """Import one workbook and return row-level results."""
    if item_type not in REQUIRED_HEADERS:
        raise ValueError('Item type must be product, reagent, or service.')

    try:
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError('The uploaded file is not a readable Excel workbook.') from exc

    if 'Upload' not in workbook.sheetnames:
        raise ValueError('The workbook must contain a sheet named "Upload".')
    sheet = workbook['Upload']
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_normalized_header(value) for value in raw_headers]
    nonblank_headers = [header for header in headers if header]
    if len(nonblank_headers) != len(set(nonblank_headers)):
        raise ValueError('The Upload sheet contains duplicate column headers.')

    missing = sorted(REQUIRED_HEADERS[item_type] - set(nonblank_headers))
    if missing:
        raise ValueError(f'Missing required columns: {", ".join(missing)}.')

    allowed_headers = set(SERVICE_HEADERS if item_type == 'service' else PRODUCT_HEADERS)
    unknown = sorted(set(nonblank_headers) - allowed_headers)
    if unknown:
        raise ValueError(f'Unknown columns: {", ".join(unknown)}.')

    parsed_rows = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {
            header: value
            for header, value in zip(headers, values)
            if header
        }
        if any(_cell_text(value) for value in row.values()):
            parsed_rows.append((row_number, row))

    if not parsed_rows:
        raise ValueError('The Upload sheet does not contain any data rows.')

    seen_ids = {}
    duplicate_rows = set()
    for row_number, row in parsed_rows:
        external_id = _cell_text(row.get('external_id')).lower()
        if external_id and external_id in seen_ids:
            duplicate_rows.add(row_number)
            duplicate_rows.add(seen_ids[external_id])
        elif external_id:
            seen_ids[external_id] = row_number

    result = {
        'item_type': item_type,
        'total_rows': len(parsed_rows),
        'created': 0,
        'updated': 0,
        'failed': 0,
        'items': [],
        'errors': [],
    }
    for row_number, row in parsed_rows:
        if row_number in duplicate_rows:
            result['failed'] += 1
            result['errors'].append({
                'row': row_number,
                'external_id': _cell_text(row.get('external_id')),
                'message': 'This External ID appears more than once in the workbook.',
            })
            continue
        try:
            with transaction.atomic():
                if item_type == 'service':
                    created, external_id, name = _import_service(row)
                else:
                    created, external_id, name = _import_product(row, item_type)
            action = 'created' if created else 'updated'
            result[action] += 1
            result['items'].append({
                'row': row_number,
                'external_id': external_id,
                'name': name,
                'action': action,
            })
        except Exception as exc:
            result['failed'] += 1
            result['errors'].append({
                'row': row_number,
                'external_id': _cell_text(row.get('external_id')),
                'message': _error_text(exc),
            })
    return result
