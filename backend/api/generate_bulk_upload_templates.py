"""Generate the three downloadable catalog bulk-upload Excel templates.

Run from the backend container so the Reference IDs sheet reflects the current
normalized Product, Reagent, and Service category/group hierarchy.
"""

import argparse
import os
import sys
import tempfile
import uuid
from pathlib import Path

import django
from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


BACKEND_DIR = Path(__file__).resolve().parents[1]
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'backend.settings')
django.setup()

from api.bulk_upload import (  # noqa: E402
    PRODUCT_HEADERS,
    REQUIRED_HEADERS,
    SERVICE_HEADERS,
    import_catalog_workbook,
)
from products.models import CatalogGroup, ProductCategory  # noqa: E402


TEMPLATES = {
    'product': 'bioark-product-bulk-upload-template.xlsx',
    'reagent': 'bioark-reagent-bulk-upload-template.xlsx',
    'service': 'bioark-service-bulk-upload-template.xlsx',
}

HEADER_HELP = {
    'external_id': 'Required. Stable public identifier. Existing IDs update the matching item.',
    'product_name': 'Required. Customer-facing product or reagent name.',
    'service_name': 'Required. Customer-facing service name (maximum 60 characters).',
    'short_description': 'Optional one-line customer-facing summary used on catalog display cards (maximum 500 characters).',
    'category_external_id': 'Required. Choose an ID from the Reference IDs sheet.',
    'group_external_id': 'Optional. Choose a group that belongs to the selected category.',
    'catalog_number': 'Optional customer-facing catalog number.',
    'show_catalog_number': 'Yes or No. Blank defaults to Yes.',
    'description': 'Plain text only. Rich-text formatting is not imported.',
    'details': 'Plain text only. Rich-text formatting is not imported.',
    'service_details': 'Plain text only. Rich-text formatting is not imported.',
    'technique': 'Plain text only. Rich-text formatting is not imported.',
    'price': 'Plain text price information for the Service Price tab.',
    'key_features': 'Plain text. Put each feature on a separate line within this cell.',
    'storage_stability': 'Plain text storage and stability information.',
    'performance_data': 'Plain text performance information.',
    'list_price': 'Base list price. Use a non-negative number, optionally prefixed with $.',
    'discounted_price': 'Optional numeric price. It cannot exceed list_price.',
    'first_option_name': 'Only this first pricing option is imported.',
    'first_option_list_price': 'Price for the first option. Requires first_option_name.',
    'first_option_discounted_price': 'Optional discounted option price; cannot exceed its list price.',
    'availability': 'Optional availability text.',
    'price_range': 'Optional price range text.',
    'quote_only': 'Yes or No. Blank defaults to No.',
    'featured': 'Yes or No. Blank defaults to No.',
    'recommended_service': 'Yes or No. Displays the service in Recommended Services.',
    'display_on_homepage': 'Yes or No. Blank defaults to No.',
    'active': 'Yes or No. Blank defaults to Yes.',
    'display_order': 'Optional whole number used to order catalog entries.',
}

BOOLEAN_HEADERS = {
    'show_catalog_number',
    'quote_only',
    'featured',
    'recommended_service',
    'display_on_homepage',
    'active',
}

WIDE_HEADERS = {
    'short_description',
    'description',
    'details',
    'service_details',
    'technique',
    'price',
    'key_features',
    'storage_stability',
    'performance_data',
}


def allowed_category_types(item_type):
    return {
        'product': {'product', 'both'},
        'reagent': {'reagent', 'consumable'},
        'service': {'service'},
    }[item_type]


def reference_records(item_type):
    allowed = allowed_category_types(item_type)
    categories = list(
        ProductCategory.objects.filter(product_type__in=allowed)
        .exclude(external_id__isnull=True)
        .exclude(external_id='')
        .order_by('priority', 'category_name', 'category_id')
    )
    groups = list(
        CatalogGroup.objects.filter(category__in=categories, is_active=True)
        .select_related('category')
        .order_by('category__priority', 'category__category_name', 'priority', 'group_name')
    )
    return categories, groups


def style_upload_sheet(sheet, item_type, headers):
    navy = '17365D'
    required_fill = PatternFill('solid', fgColor='F4B183')
    optional_fill = PatternFill('solid', fgColor='5B9BD5')
    white_font = Font(color='FFFFFF', bold=True, size=10)
    required_font = Font(color='172033', bold=True, size=10)
    thin = Side(style='thin', color='D7E0EA')

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1001'
    sheet.row_dimensions[1].height = 38

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        required = header in REQUIRED_HEADERS[item_type]
        cell.fill = required_fill if required else optional_fill
        cell.font = required_font if required else white_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        cell.comment = Comment(HEADER_HELP.get(header, ''), 'BioArk Technologies')

        if header in WIDE_HEADERS:
            width = 38
        elif header in {'product_name', 'service_name'}:
            width = 32
        elif header in {'external_id', 'category_external_id', 'group_external_id'}:
            width = 29
        elif header in BOOLEAN_HEADERS:
            width = 22
        elif 'price' in header:
            width = 24
        else:
            width = 21
        sheet.column_dimensions[get_column_letter(column_index)].width = width

        if header in BOOLEAN_HEADERS:
            validation = DataValidation(
                type='list',
                formula1='"Yes,No"',
                allow_blank=True,
                error='Choose Yes or No.',
                errorTitle='Invalid selection',
                prompt='Choose Yes or No. Blank uses the documented default.',
                promptTitle=header,
                showErrorMessage=True,
                showInputMessage=True,
            )
            sheet.add_data_validation(validation)
            validation.add(f'{get_column_letter(column_index)}2:{get_column_letter(column_index)}1001')
        elif header == 'category_external_id':
            validation = DataValidation(
                type='list',
                formula1='=CategoryIDs',
                allow_blank=False,
                error='Choose a Category External ID from the Reference IDs sheet.',
                errorTitle='Unknown category',
                showErrorMessage=True,
            )
            sheet.add_data_validation(validation)
            validation.add(f'{get_column_letter(column_index)}2:{get_column_letter(column_index)}1001')
        elif header == 'group_external_id':
            validation = DataValidation(
                type='list',
                formula1='=GroupIDs',
                allow_blank=True,
                error='Choose a Group External ID from the Reference IDs sheet.',
                errorTitle='Unknown group',
                showErrorMessage=True,
            )
            sheet.add_data_validation(validation)
            validation.add(f'{get_column_letter(column_index)}2:{get_column_letter(column_index)}1001')

    sheet.conditional_formatting.add(
        f'A2:{get_column_letter(len(headers))}1001',
        FormulaRule(formula=['MOD(ROW(),2)=0'], fill=PatternFill('solid', fgColor='F7FAFD')),
    )
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.tabColor = navy


def add_instruction_sheet(workbook, item_type, headers):
    sheet = workbook.create_sheet('Instructions')
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions['A'].width = 4
    sheet.column_dimensions['B'].width = 29
    sheet.column_dimensions['C'].width = 92
    sheet.merge_cells('B2:C2')
    sheet['B2'] = f'BioArk {item_type.title()} Bulk Upload Template'
    sheet['B2'].font = Font(size=20, bold=True, color='17365D')
    sheet['B2'].alignment = Alignment(vertical='center')
    sheet.row_dimensions[2].height = 34

    rules = [
        'Enter one catalog item per row on the Upload sheet. Do not add title or instruction rows.',
        'Use the matching Product, Reagent, or Service template. Do not combine item types.',
        'Required columns have orange headers. Optional columns have blue headers.',
        'External ID is the stable match key: an existing ID updates that item; a new ID creates it.',
        'Text is imported as plain text. Rich-text formatting from Excel is ignored.',
        'Images, videos, documents, and manuals are not imported or removed.',
        'Products and Reagents import only first_option_name and its two price fields.',
        'Blank optional cells clear that imported field on an existing item; media remains unchanged.',
        'Use category and group IDs from the Reference IDs sheet. The selected group must belong to the category.',
        'Save as .xlsx and keep the Upload sheet name and column headers unchanged.',
    ]
    sheet['B4'] = 'How to use this template'
    sheet['B4'].font = Font(size=13, bold=True, color='17365D')
    for index, rule in enumerate(rules, start=5):
        sheet[f'B{index}'] = index - 4
        sheet[f'B{index}'].font = Font(bold=True, color='2F75B5')
        sheet[f'C{index}'] = rule
        sheet[f'C{index}'].alignment = Alignment(wrap_text=True, vertical='top')
        sheet.row_dimensions[index].height = 30

    start = 17
    sheet[f'B{start}'] = 'Column'
    sheet[f'C{start}'] = 'Meaning'
    for cell in sheet[f'B{start}:C{start}'][0]:
        cell.fill = PatternFill('solid', fgColor='17365D')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(vertical='center')
    for row_index, header in enumerate(headers, start=start + 1):
        sheet[f'B{row_index}'] = header
        sheet[f'B{row_index}'].font = Font(bold=header in REQUIRED_HEADERS[item_type])
        sheet[f'C{row_index}'] = HEADER_HELP.get(header, '')
        sheet[f'C{row_index}'].alignment = Alignment(wrap_text=True, vertical='top')
        fill = 'FFF2CC' if header in REQUIRED_HEADERS[item_type] else ('F6F9FC' if row_index % 2 == 0 else 'FFFFFF')
        sheet[f'B{row_index}'].fill = PatternFill('solid', fgColor=fill)
        sheet[f'C{row_index}'].fill = PatternFill('solid', fgColor=fill)
        sheet.row_dimensions[row_index].height = 28
    sheet.freeze_panes = f'B{start + 1}'
    sheet.sheet_properties.tabColor = '70AD47'


def add_reference_sheet(workbook, categories, groups):
    sheet = workbook.create_sheet('Reference IDs')
    sheet.sheet_view.showGridLines = False
    category_headers = ['Category External ID', 'Category Name']
    group_headers = ['Group External ID', 'Group Name', 'Category External ID']
    sheet['A1'] = category_headers[0]
    sheet['B1'] = category_headers[1]
    sheet['D1'] = group_headers[0]
    sheet['E1'] = group_headers[1]
    sheet['F1'] = group_headers[2]
    for cell in sheet['A1:B1'][0] + sheet['D1:F1'][0]:
        cell.fill = PatternFill('solid', fgColor='17365D')
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row_index, category in enumerate(categories, start=2):
        sheet.cell(row=row_index, column=1, value=category.external_id)
        sheet.cell(row=row_index, column=2, value=category.category_name)
    for row_index, group in enumerate(groups, start=2):
        sheet.cell(row=row_index, column=4, value=group.external_id)
        sheet.cell(row=row_index, column=5, value=group.group_name)
        sheet.cell(row=row_index, column=6, value=group.category.external_id)

    for column, width in {'A': 31, 'B': 34, 'C': 4, 'D': 34, 'E': 34, 'F': 31}.items():
        sheet.column_dimensions[column].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:F{max(len(categories), len(groups), 1) + 1}'
    sheet.sheet_properties.tabColor = 'A5A5A5'

    category_end = max(2, len(categories) + 1)
    group_end = max(2, len(groups) + 1)
    workbook.defined_names.add(DefinedName(
        'CategoryIDs', attr_text=f"'Reference IDs'!$A$2:$A${category_end}"
    ))
    workbook.defined_names.add(DefinedName(
        'GroupIDs', attr_text=f"'Reference IDs'!$D$2:$D${group_end}"
    ))


def create_template(item_type, output_path):
    headers = SERVICE_HEADERS if item_type == 'service' else PRODUCT_HEADERS
    categories, groups = reference_records(item_type)
    if not categories:
        raise RuntimeError(f'No valid {item_type} categories were found for the reference sheet.')

    workbook = Workbook()
    upload = workbook.active
    upload.title = 'Upload'
    for column_index, header in enumerate(headers, start=1):
        upload.cell(row=1, column=column_index, value=header)
    add_instruction_sheet(workbook, item_type, headers)
    add_reference_sheet(workbook, categories, groups)
    style_upload_sheet(upload, item_type, headers)
    workbook.active = 0
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def verify_template(item_type, path):
    expected_headers = SERVICE_HEADERS if item_type == 'service' else PRODUCT_HEADERS
    workbook = load_workbook(path)
    assert workbook.sheetnames == ['Upload', 'Instructions', 'Reference IDs']
    upload = workbook['Upload']
    headers = [upload.cell(row=1, column=index).value for index in range(1, len(expected_headers) + 1)]
    assert headers == expected_headers
    assert REQUIRED_HEADERS[item_type].issubset(set(headers))
    assert not ({'images', 'image', 'documents', 'manuals', 'videos'} & set(headers))
    assert upload.freeze_panes == 'A2'
    assert len(upload.data_validations.dataValidation) >= 3
    assert workbook['Reference IDs'].max_row >= 2
    assert 'CategoryIDs' in workbook.defined_names
    assert 'GroupIDs' in workbook.defined_names

    category_id = workbook['Reference IDs']['A2'].value
    group_id = ''
    for row in workbook['Reference IDs'].iter_rows(min_row=2, min_col=4, max_col=6, values_only=True):
        if row[0] and row[2] == category_id:
            group_id = row[0]
            break

    verification_id = f'template-check-{item_type}-{uuid.uuid4().hex[:8]}'
    values = {
        'external_id': verification_id,
        'category_external_id': category_id,
        'group_external_id': group_id,
        'active': 'Yes',
        'show_catalog_number': 'Yes',
        'short_description': 'Template verification short description.',
    }
    if item_type == 'service':
        values.update({'service_name': 'Template Verification Service', 'service_details': 'Plain text'})
    else:
        values.update({
            'product_name': f'Template Verification {item_type.title()}',
            'first_option_name': 'First option',
            'first_option_list_price': '19.95',
        })

    with tempfile.TemporaryDirectory() as temporary_directory:
        test_path = Path(temporary_directory) / path.name
        for column_index, header in enumerate(headers, start=1):
            upload.cell(row=2, column=column_index, value=values.get(header, ''))
        workbook.save(test_path)
        with transaction.atomic():
            with test_path.open('rb') as handle:
                result = import_catalog_workbook(handle, item_type)
            if result['created'] != 1 or result['failed'] != 0:
                raise AssertionError(f'{item_type} parser verification failed: {result}')
            transaction.set_rollback(True)
    return {
        'item_type': item_type,
        'file': path.name,
        'headers': len(headers),
        'categories': sum(
            1 for cell in workbook['Reference IDs']['A'][1:] if cell.value
        ),
        'groups': sum(
            1 for cell in workbook['Reference IDs']['D'][1:] if cell.value
        ),
        'data_validations': len(upload.data_validations.dataValidation),
        'parser_created': 1,
        'parser_failed': 0,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('output_directory', type=Path)
    args = parser.parse_args()

    reports = []
    for item_type, filename in TEMPLATES.items():
        output_path = args.output_directory / filename
        create_template(item_type, output_path)
        reports.append(verify_template(item_type, output_path))
    for report in reports:
        print(report)


if __name__ == '__main__':
    main()
